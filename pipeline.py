#!/usr/bin/env python3
"""
Pipeline: Konversi 4 file Excel Master Aset BMN → Satu file Parquet terkompresi.
Optimasi RAM: Membaca file satu per satu via openpyxl read-only streaming,
memproses dalam chunk 50K baris, menulis langsung ke disk.
"""

import os
import sys
import gc
from pathlib import Path
from datetime import datetime

import openpyxl
import polars as pl

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input_data"
CACHE_DIR = BASE_DIR / "data_cache"

GOLONGAN_MAP = {
    "1": "PERSEDIAAN",
    "2": "TANAH",
    "3": "PERALATAN DAN MESIN",
    "4": "GEDUNG DAN BANGUNAN",
    "5": "JALAN, IRIGASI DAN JARINGAN",
    "6": "ASET TETAP LAINNYA",
    "7": "KONSTRUKSI DALAM PENGERJAAN",
    "8": "ASET TAK BERWUJUD",
}

KOLOM_KEYS = {
    "no": "nomor",
    "jenis bmn": "jenis_bmn",
    "jenis_bmn": "jenis_bmn",
    "kode satker": "kode_satker",
    "kode_satker": "kode_satker",
    "nama satker": "nama_satker",
    "nama_satker": "nama_satker",
    "kode barang": "kode_barang",
    "kode_barang": "kode_barang",
    "nup": "nup",
    "nama barang": "nama_barang",
    "status bmn": "status_bmn",
    "merk": "merk",
    "tipe": "tipe",
    "kondisi": "kondisi",
    "umur aset": "umur_aset",
    "intra / extra": "intra_extra",
    "henti guna": "henti_guna",
    "status sbsn": "status_sbsn",
    "status bmn idle": "status_bmn_idle",
    "status kemitraan": "status_kemitraan",
    "bpybds": "bpybds",
    "usulan barang hilang": "usulan_barang_hilang",
    "usulan barang rb": "usulan_barang_rb",
    "usul hapus": "usul_hapus",
    "hibah dktp": "hibah_dktp",
    "konsensi jasa": "konsensi_jasa",
    "properti investasi": "properti_investasi",
    "jenis dokumen": "jenis_dokumen",
    "no dokumen": "no_dokumen",
    "no bpkp": "no_bpkp",
    "no polisi": "no_polisi",
    "status sertifikasi": "status_sertifikasi",
    "jenis sertipikat": "jenis_sertipikat",
    "no sertifikat": "no_sertifikat",
    "nama": "nama",
    "tanggal buku pertama": "tanggal_buku_pertama",
    "tanggal perolehan": "tanggal_perolehan",
    "tanggal pengapusan": "tanggal_pengapusan",
    "nilai perolehan pertama": "nilai_perolehan_pertama",
    "nilai mutasi": "nilai_mutasi",
    "nilai perolehan": "nilai_perolehan",
    "nilai penyusutan": "nilai_penyusutan",
    "nilai buku": "nilai_buku",
    "luas tanah seluruhnya": "luas_tanah_seluruhnya",
    "luas tanah untuk bangunan": "luas_tanah_bangunan",
    "luas tanah untuk sarana lingkungan": "luas_tanah_sarana",
    "luas lahan kosong": "luas_lahan_kosong",
    "luas bangunan": "luas_bangunan",
    "luas tapak bangunan": "luas_tapak_bangunan",
    "luas pemanfataan": "luas_pemanfaatan",
    "jumlah lantai": "jumlah_lantai",
    "jumlah foto": "jumlah_foto",
    "status penggunaan": "status_penggunaan",
    "no psp": "no_psp",
    "tanggal psp": "tanggal_psp",
    "alamat": "alamat",
    "rt/rw": "rt_rw",
    "kelurahan/desa": "kelurahan_desa",
    "kecamatan": "kecamatan",
    "kab/kota": "kab_kota",
    "kode kab/kota": "kode_kab_kota",
    "provinsi": "provinsi",
    "kode provinsi": "kode_provinsi",
    "kode pos": "kode_pos",
    "sbsk": "sbsk",
    "optimalisasi": "optimalisasi",
    "penghuni": "penghuni",
    "pengguna": "pengguna",
    "kode kpknl": "kode_kpknl",
    "uraian kpknl": "uraian_kpknl",
    "uraian kanwil djkn": "uraian_kanwil_djkn",
    "nama k/l": "nama_kl",
    "nama e1": "nama_e1",
    "nama korwil": "nama_korwil",
    "kode register": "kode_register",
    "lokasi ruang": "lokasi_ruang",
    "jenis identitas": "jenis_identitas",
    "no identitas": "no_identitas",
    "no stnk": "no_stnk",
    "nama pengguna": "nama_pengguna",
    "status pmk": "status_pmk",
}

COLUMN_ORDER = [
    "nomor", "jenis_bmn", "kode_satker", "nama_satker", "kode_barang", "nup",
    "golongan_bmn", "nama_barang", "merk", "tipe", "kondisi", "umur_aset",
    "nilai_perolehan", "nilai_buku", "nilai_penyusutan", "nilai_perolehan_pertama", "nilai_mutasi",
    "tanggal_perolehan", "tanggal_buku_pertama", "tanggal_pengapusan",
    "alamat", "rt_rw", "kelurahan_desa", "kecamatan", "kab_kota", "provinsi", "kode_pos",
    "status_penggunaan", "status_bmn", "intra_extra",
    "no_dokumen", "no_sertifikat", "no_psp", "no_polisi",
    "luas_tanah_seluruhnya", "luas_bangunan", "jumlah_foto", "jumlah_lantai",
    "sumber_file",
]

KOLOM_NUMERIK = {
    "nomor", "nup", "umur_aset",
    "nilai_perolehan", "nilai_buku", "nilai_penyusutan",
    "nilai_perolehan_pertama", "nilai_mutasi",
    "luas_tanah_seluruhnya", "luas_tanah_bangunan", "luas_tanah_sarana", "luas_lahan_kosong",
    "luas_bangunan", "luas_tapak_bangunan", "luas_pemanfaatan",
    "jumlah_lantai", "jumlah_foto",
}

KOLOM_TANGGAL = {
    "tanggal_perolehan", "tanggal_buku_pertama", "tanggal_pengapusan", "tanggal_psp",
}


def normalisasi_kolom(df: pl.DataFrame) -> pl.DataFrame:
    mapping = {}
    for col in df.columns:
        key = col.strip().lower().replace("  ", " ")
        if key in KOLOM_KEYS:
            target = KOLOM_KEYS[key]
            if target != col:
                mapping[col] = target
    if mapping:
        df = df.rename(mapping)
    return df


def cari_file_excel(base_dir: Path) -> list[Path]:
    if INPUT_DIR.exists():
        files = sorted(INPUT_DIR.glob("*.xlsx"))
        if files:
            return files
    files = sorted(base_dir.parent.glob("*.xlsx"))
    return files


def baca_excel_chunking(filepath: Path, chunk_size: int = 50_000) -> pl.DataFrame:
    MAX_ROW = 99_999_999
    MAX_COL = 100

    print(f"  Membuka: {filepath.name} ({filepath.stat().st_size / 1024 / 1024:.1f} MB)")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.title}")

    header_row = next(ws.iter_rows(min_row=1, max_row=MAX_ROW, min_col=1, max_col=MAX_COL, values_only=True))
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

    import datetime as _dt
    def _serialize(val):
        if val is None:
            return None
        if isinstance(val, _dt.datetime):
            return val.strftime("%Y-%m-%d")
        if isinstance(val, _dt.date):
            return val.isoformat()
        return str(val)

    chunks = []
    batch = []
    processed = 0
    skipped_empty = 0

    for row in ws.iter_rows(min_row=2, max_row=MAX_ROW, min_col=1, max_col=MAX_COL, values_only=True):
        vals = [_serialize(v) for v in row]
        if all(v is None for v in vals):
            skipped_empty += 1
            continue

        batch.append(vals)
        processed += 1

        if len(batch) >= chunk_size:
            df_chunk = pl.DataFrame(batch, schema=headers, orient="row", infer_schema_length=None)
            chunks.append(df_chunk)
            batch = []
            print(f"    → {processed:,} baris terbaca...")
            gc.collect()

    if batch:
        df_chunk = pl.DataFrame(batch, schema=headers, orient="row", infer_schema_length=None)
        chunks.append(df_chunk)

    wb.close()

    if skipped_empty > 0:
        print(f"    ℹ {skipped_empty} baris kosong dilewati")

    if not chunks:
        print("    ⚠ Tidak ada data.")
        return pl.DataFrame(schema=headers)

    print(f"    → Menggabungkan {len(chunks)} chunk...")
    df = pl.concat(chunks, how="vertical")

    all_none_cols = [c for c in df.columns if df.select(pl.col(c).is_null().all()).item()]
    if all_none_cols:
        df = df.drop(all_none_cols)
        print(f"    ℹ {len(all_none_cols)} kolom kosong dihapus")

    print(f"    → Total: {df.height:,} baris, {df.width} kolom")
    return df


def pipeline():
    print("=" * 60)
    print("PIPELINE KONVERSI EXCEL → PARQUET")
    print(f"Waktu mulai: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    files = cari_file_excel(BASE_DIR)
    if not files:
        print("Tidak ditemukan file Excel (.xlsx) di folder input_data/ atau root proyek.")
        print("Silakan letakkan 4 file Excel di folder input_data/")
        sys.exit(1)

    print(f"\nDitemukan {len(files)} file Excel:")
    for f in files:
        print(f"   - {f.name} ({f.stat().st_size / 1024 / 1024:.1f} MB)")
    print()

    semua_dfs = []
    total_rows = 0

    for idx, filepath in enumerate(files, 1):
        print(f"[{idx}/{len(files)}] Proses file...")
        df = baca_excel_chunking(filepath)

        if df.height == 0:
            print("    File kosong, dilewati.\n")
            continue

        df = normalisasi_kolom(df)
        df = df.with_columns(
            [
                pl.lit(filepath.stem).alias("sumber_file"),
                pl.col("kode_barang").str.slice(0, 1).replace(GOLONGAN_MAP, default="TIDAK DIKENAL").alias("golongan_bmn"),
            ]
        )

        for col in COLUMN_ORDER:
            if col not in df.columns:
                df = df.with_columns(pl.lit(None).alias(col))

        cols_lain = [c for c in df.columns if c not in COLUMN_ORDER]
        df = df.select(COLUMN_ORDER + cols_lain)

        for col in df.columns:
            if col in KOLOM_NUMERIK:
                df = df.with_columns(pl.col(col).cast(pl.Float64, strict=False))
            elif col in KOLOM_TANGGAL:
                df = df.with_columns(pl.col(col).cast(pl.Date, strict=False))
            elif col in ("nomor", "nup"):
                df = df.with_columns(pl.col(col).cast(pl.Int64, strict=False))
            else:
                df = df.with_columns(pl.col(col).cast(pl.Utf8, strict=False))

        semua_dfs.append(df)
        total_rows += df.height
        print(f"    Selesai memproses {df.height:,} baris\n")

    print("-" * 60)
    print(f"Menggabungkan {len(semua_dfs)} DataFrame...")
    df_gabungan = pl.concat(semua_dfs, how="vertical")
    print(f"   Total baris: {df_gabungan.height:,}")
    print(f"   Total kolom: {df_gabungan.width}")
    size_mb = df_gabungan.estimated_size("mb")
    print(f"   Estimasi memori: {size_mb:.1f} MB")

    parquet_path = CACHE_DIR / "master_aset_bali.parquet"
    print(f"\nMenulis file Parquet: {parquet_path}")
    df_gabungan.write_parquet(parquet_path, compression="snappy", statistics=True)
    size_mb = parquet_path.stat().st_size / 1024 / 1024
    print(f"   Tersimpan: {size_mb:.1f} MB")

    print("\n" + "=" * 60)
    print("PIPELINE SELESAI")
    print(f"   File output: {parquet_path}")
    print(f"   Total baris: {df_gabungan.height:,}")
    print(f"   Waktu selesai: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    return df_gabungan


if __name__ == "__main__":
    pipeline()
